"""Lectura del workbook PMO TI mensual.

Encapsula la lógica para extraer la lista canónica de proyectos del
Índice y, por cada hoja, las celdas estándar que alimentan el PPT.

Uso típico:

    from openpyxl import load_workbook
    from scripts.excel_reader import leer_indice, leer_proyecto

    wb = load_workbook(path, data_only=True)
    proyectos = leer_indice(wb)
    for sheet_name in proyectos:
        data = leer_proyecto(wb[sheet_name])
        # data['nombre'], data['pm_ti'], data['hitos'], data['notas'], data['riesgos']…
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Iterator

from .helpers import clean_section_title, extract_sheet_name_from_formula


SHEETS_NO_PROYECTO = {"Instrucciones", "Índice", "Indice", "Plantilla"}


@dataclass
class Hito:
    ref: str            # H-01, H-02…
    entregable: str
    inicio: date | datetime | str | None
    fin: date | datetime | str | None
    estado: str         # COMPLETADO / EN CURSO / PENDIENTE / BLOQUEADO / PLANIFICADO
    comentario: str = ""


@dataclass
class Nota:
    icono: str          # ✓ / ⚠️ / 📓
    texto: str


@dataclass
class Riesgo:
    ref: str            # R-01, R-02…
    descripcion: str
    impacto: str        # ALTO / MEDIO / BAJO
    plan: str
    responsable: str
    deadline: date | datetime | str | None


@dataclass
class Proyecto:
    sheet_name: str                  # nombre real de la pestaña
    nombre: str                      # B3 — fuente de verdad para el PPT
    direccion: str = ""              # F6
    pm_ti: str = ""                  # C7
    patrocinador: str = ""           # F7
    pm_pmo: str = ""                 # C8
    fecha_inicio: date | datetime | str | None = None  # F8
    fecha_reporte: date | datetime | str | None = None # C9
    fecha_fin: date | datetime | str | None = None     # F9
    lider_tecnico: str = ""          # B10/C10
    seguimiento: str = ""            # F10 — ON TRACK / EN RIESGO / OFF TRACK
    pct_avance: float = 0.0          # C11
    pct_desviacion: float = 0.0      # F11
    alcance: str = ""                # C12
    titulo_entregables: str = "ENTREGABLES"          # parsed del prefijo "2. "
    titulo_notas: str = "NOTAS RELEVANTES"           # parsed del prefijo "3. "
    titulo_riesgos: str = "RIESGOS Y PLAN DE MITIGACIÓN"  # parsed del prefijo "4. "
    hitos: list[Hito] = field(default_factory=list)
    notas: list[Nota] = field(default_factory=list)
    riesgos: list[Riesgo] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Índice
# ---------------------------------------------------------------------------

def leer_indice(wb) -> list[str]:
    """Devuelve la lista de sheet_names de los proyectos activos, en orden.

    Lee la pestaña `Índice` y extrae el sheet name de cada proyecto
    parseando la fórmula `='SheetName'!B3` de la columna correspondiente.
    Si no encuentra la pestaña o las fórmulas no parsean, cae a iterar
    sobre todas las pestañas excluyendo las no-proyecto.
    """
    nombre_indice = "Índice" if "Índice" in wb.sheetnames else ("Indice" if "Indice" in wb.sheetnames else None)
    if nombre_indice is None:
        return [s for s in wb.sheetnames if s not in SHEETS_NO_PROYECTO and not s.startswith("Ejemplo")]

    ws = wb[nombre_indice]
    proyectos: list[str] = []
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            val = cell.value
            sheet = extract_sheet_name_from_formula(val)
            if sheet and sheet not in SHEETS_NO_PROYECTO and sheet in wb.sheetnames:
                proyectos.append(sheet)
                break  # una vez encontrado el sheet de esta fila, pasar a la siguiente
    # Si parsing falla totalmente, fallback
    if not proyectos:
        proyectos = [s for s in wb.sheetnames if s not in SHEETS_NO_PROYECTO and not s.startswith("Ejemplo")]
    return proyectos


# ---------------------------------------------------------------------------
# Hoja de proyecto
# ---------------------------------------------------------------------------

_HITO_RE = re.compile(r"^H-\d+", re.IGNORECASE)
_RIESGO_RE = re.compile(r"^R-\d+", re.IGNORECASE)


def _cell(ws, ref: str):
    v = ws[ref].value
    return v if v is not None else ""


def _find_section_row(ws, prefix: str) -> int | None:
    """Busca la primera fila de la columna B cuyo valor empieza con el prefijo dado.

    `prefix` debería ser `"2."`, `"3."`, `"4."` (con o sin espacio).
    """
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=2).value  # columna B
        if isinstance(v, str) and v.strip().startswith(prefix):
            return row
    return None


def leer_proyecto(ws) -> Proyecto:
    """Extrae todos los datos de una hoja de proyecto a un dataclass Proyecto."""
    p = Proyecto(
        sheet_name=ws.title,
        nombre=str(_cell(ws, "B3")).strip() or "[SIN TÍTULO — REVISAR B3]",
        direccion=str(_cell(ws, "F6")).strip(),
        pm_ti=str(_cell(ws, "C7")).strip(),
        patrocinador=str(_cell(ws, "F7")).strip(),
        pm_pmo=str(_cell(ws, "C8")).strip(),
        fecha_inicio=_cell(ws, "F8") or None,
        fecha_reporte=_cell(ws, "C9") or None,
        fecha_fin=_cell(ws, "F9") or None,
        lider_tecnico=str(_cell(ws, "B10") or _cell(ws, "C10")).strip(),
        seguimiento=str(_cell(ws, "F10")).strip().upper(),
        pct_avance=_safe_pct(_cell(ws, "C11")),
        pct_desviacion=_safe_pct(_cell(ws, "F11")),
        alcance=str(_cell(ws, "C12")).strip(),
    )

    # Títulos parseados
    row2 = _find_section_row(ws, "2.")
    row3 = _find_section_row(ws, "3.")
    row4 = _find_section_row(ws, "4.")

    if row2 is not None:
        p.titulo_entregables = clean_section_title(ws.cell(row=row2, column=2).value) or p.titulo_entregables
    if row3 is not None:
        p.titulo_notas = clean_section_title(ws.cell(row=row3, column=2).value) or p.titulo_notas
    if row4 is not None:
        p.titulo_riesgos = clean_section_title(ws.cell(row=row4, column=2).value) or p.titulo_riesgos

    # Hitos: filas H-NN entre row2+1 (header) y row3-1
    if row2 is not None and row3 is not None:
        for row in range(row2 + 2, row3):  # +2 salta título y header
            ref = str(ws.cell(row=row, column=2).value or "").strip()
            if not _HITO_RE.match(ref):
                continue
            entregable = str(ws.cell(row=row, column=3).value or "").strip()
            if not entregable:
                continue  # filas vacías se omiten
            p.hitos.append(Hito(
                ref=ref,
                entregable=entregable,
                inicio=ws.cell(row=row, column=4).value,
                fin=ws.cell(row=row, column=5).value,
                estado=str(ws.cell(row=row, column=6).value or "").strip().upper(),
                comentario=str(ws.cell(row=row, column=7).value or "").strip(),
            ))

    # Notas: hasta 4 filas tras row3 (saltando la leyenda)
    if row3 is not None and row4 is not None:
        for row in range(row3 + 1, row4):
            icono = str(ws.cell(row=row, column=2).value or "").strip()
            texto = str(ws.cell(row=row, column=3).value or "").strip()
            # Saltar la leyenda
            if texto.lower().startswith("leyenda"):
                continue
            if texto and icono:
                p.notas.append(Nota(icono=icono, texto=texto))
            if len(p.notas) >= 4:
                break

    # Riesgos: filas R-NN tras row4
    if row4 is not None:
        for row in range(row4 + 2, ws.max_row + 1):
            ref = str(ws.cell(row=row, column=2).value or "").strip()
            if not _RIESGO_RE.match(ref):
                # detener si encontramos el footer u otra cosa
                if ref.lower().startswith("versión") or ref == "":
                    continue
                break
            descripcion = str(ws.cell(row=row, column=3).value or "").strip()
            if not descripcion:
                continue
            p.riesgos.append(Riesgo(
                ref=ref,
                descripcion=descripcion,
                impacto=str(ws.cell(row=row, column=4).value or "").strip().upper(),
                plan=str(ws.cell(row=row, column=5).value or "").strip(),
                responsable=str(ws.cell(row=row, column=6).value or "").strip(),
                deadline=ws.cell(row=row, column=7).value,
            ))

    return p


def _safe_pct(v) -> float:
    """Normaliza un % que puede venir como 0.85, 85, '85%', etc. a un float en [0,100]."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        f = float(v)
        return f * 100 if 0 < f <= 1 else f
    s = str(v).replace("%", "").strip()
    try:
        f = float(s)
        return f * 100 if 0 < f <= 1 else f
    except ValueError:
        return 0.0


def iter_proyectos(wb) -> Iterator[Proyecto]:
    """Itera los proyectos del Índice en orden, ya parseados."""
    for sheet_name in leer_indice(wb):
        yield leer_proyecto(wb[sheet_name])
